from decimal import Decimal
from io import BytesIO

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .models import TransactionType


HEADER_FILL = PatternFill('solid', fgColor='1E293B')
HEADER_FONT = Font(bold=True, color='FFFFFF')
MONEY_FMT = '#,##0.00'


def _style_header_row(ws, row=1):
  for cell in ws[row]:
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(horizontal='center', vertical='center')


def _autosize_columns(ws, max_width=48):
  for col_idx, column_cells in enumerate(ws.columns, start=1):
    length = 0
    for cell in column_cells:
      if cell.value is not None:
        length = max(length, len(str(cell.value)))
    ws.column_dimensions[get_column_letter(col_idx)].width = min(length + 2, max_width)


def _write_summary_sheet(ws, summary):
  ws.title = 'Сводка'
  rows = [
    ('Период', summary['period']),
    ('Доходы', float(summary['income'])),
    ('Расходы', float(summary['expense'])),
    ('Итого (доходы − расходы)', float(summary['net'])),
    ('Погашения задолженностей', float(summary['debt_repayments'])),
    ('Баланс (общий)', float(summary['main_balance'])),
    ('Резерв', float(summary['reserve_balance'])),
    ('Коммунальные', float(summary['utilities_balance'])),
    ('Транзакций в отчёте', summary['transactions_count']),
    ('Операций резерва', summary['reserve_count']),
    ('Операций коммунальных', summary['utilities_count']),
  ]
  for row_idx, (label, value) in enumerate(rows, start=1):
    ws.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
    cell = ws.cell(row=row_idx, column=2, value=value)
    if isinstance(value, float):
      cell.number_format = MONEY_FMT
  ws.column_dimensions['A'].width = 32
  ws.column_dimensions['B'].width = 20


def _write_transactions_sheet(ws, transactions):
  ws.title = 'Транзакции'
  headers = [
    'Дата', 'Название', 'Тип', 'Проект', 'Задолженность',
    'Должник', 'Сумма', 'На коммунальные', 'На долг', 'Заметки',
  ]
  ws.append(headers)
  _style_header_row(ws)

  for t in transactions:
    ws.append([
      t.date,
      t.title,
      t.get_transaction_type_display(),
      t.project.name if t.project_id else '',
      t.debt.title if t.debt_id else '',
      t.repaid_by_display if t.transaction_type == TransactionType.DEBT_REPAYMENT else '',
      float(t.amount),
      float(t.utilities_portion) if t.transaction_type == TransactionType.DEBT_REPAYMENT else '',
      float(t.debt_reduction_amount) if t.transaction_type == TransactionType.DEBT_REPAYMENT else '',
      t.notes or '',
    ])

  for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    row[6].number_format = MONEY_FMT
    if row[7].value not in ('', None):
      row[7].number_format = MONEY_FMT
    if row[8].value not in ('', None):
      row[8].number_format = MONEY_FMT

  _autosize_columns(ws)


def _write_reserve_sheet(ws, transfers):
  ws.title = 'Резерв'
  headers = ['Дата', 'Название', 'Операция', 'Источник', 'Проект', 'Сумма', 'Заметки']
  ws.append(headers)
  _style_header_row(ws)
  for item in transfers:
    ws.append([
      item.date,
      item.title,
      item.get_direction_display(),
      item.source_display,
      item.project.name if item.project_id else '',
      float(item.amount),
      item.notes or '',
    ])
  for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    row[5].number_format = MONEY_FMT
  _autosize_columns(ws)


def _write_utilities_sheet(ws, operations):
  ws.title = 'Коммунальные'
  headers = ['Дата', 'Название', 'Тип', 'Источник', 'Проект', 'Задолженность', 'Сумма', 'Заметки']
  ws.append(headers)
  _style_header_row(ws)
  for op in operations:
    ws.append([
      op.date,
      op.title,
      op.get_operation_type_display(),
      op.source_display,
      op.project.name if op.project_id else '',
      op.debt.title if op.debt_id else '',
      float(op.amount),
      op.notes or '',
    ])
  for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    row[6].number_format = MONEY_FMT
  _autosize_columns(ws)


def build_finance_excel(summary, transactions, reserve_transfers, utilities_ops):
  wb = Workbook()
  _write_summary_sheet(wb.active, summary)

  ws_tx = wb.create_sheet()
  _write_transactions_sheet(ws_tx, transactions)

  ws_res = wb.create_sheet()
  _write_reserve_sheet(ws_res, reserve_transfers)

  ws_util = wb.create_sheet()
  _write_utilities_sheet(ws_util, utilities_ops)

  buffer = BytesIO()
  wb.save(buffer)
  buffer.seek(0)
  return buffer


def finance_excel_response(summary, transactions, reserve_transfers, utilities_ops, year, month, all_time):
  buffer = build_finance_excel(summary, transactions, reserve_transfers, utilities_ops)
  if all_time:
    filename = 'finance_report_all.xlsx'
  else:
    filename = f'finance_report_{year:04d}-{month:02d}.xlsx'
  response = HttpResponse(
    buffer.getvalue(),
    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
  )
  response['Content-Disposition'] = f'attachment; filename="{filename}"'
  return response
